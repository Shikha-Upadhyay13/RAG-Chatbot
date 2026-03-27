from typing import List, Dict
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from fastapi import UploadFile


def extract_text_from_file(file: UploadFile) -> List[Dict]:
    """
    Extract text from file WITH metadata.

    Returns a list of:
    {
        "text": str,
        "page": int | None,
        "source": filename
    }
    """

    filename = file.filename
    lower = filename.lower()

    documents: List[Dict] = []

    # -------- PDF --------
    if lower.endswith(".pdf"):
        reader = PdfReader(file.file)

        for page_idx, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            text = text.strip()

            if not text:
                continue

            documents.append({
                "text": text,
                "page": page_idx + 1,
                "source": filename,
            })

        return documents

    # -------- DOCX / DOC --------
    if lower.endswith(".docx") or lower.endswith(".doc"):
        doc = Document(file.file)
        full_text = "\n".join(p.text for p in doc.paragraphs).strip()

        if not full_text:
            return []

        documents.append({
            "text": full_text,
            "page": None,
            "source": filename,
        })

        return documents

    # -------- EXCEL (.xlsx, .xls) --------
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        workbook = load_workbook(file.file, read_only=True, data_only=True)

        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            ws = workbook[sheet_name]
            rows = []

            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                line = " | ".join(cells)
                if line.strip():
                    rows.append(line)

            sheet_text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            sheet_text = sheet_text.strip()

            if not sheet_text or sheet_text == f"Sheet: {sheet_name}":
                continue

            documents.append({
                "text": sheet_text,
                "page": sheet_idx + 1,
                "source": filename,
            })

        workbook.close()
        return documents

    # -------- UNSUPPORTED --------
    raise ValueError(f"Unsupported file type: {filename}. Supported: PDF, DOCX, DOC, XLSX, XLS")
